import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def f(x, y):
    # return y * (x - 1) - (2 - 0.5 * x) * x
    return x ** 2 + y ** 2 - 5 * x * y


# A method for calculating the solution of a differential equation by the Euler method
# The method receives a function f, x0 (initial value of x), y0 (initial value of y), h (step) and xmax
# (the maximum value of x up to which the calculation will be performed)
def euler_method(f, x0, y0, h, xmax):
    # Calculate the number of steps
    n = int((xmax - x0) / h) + 1

    # Initialize arrays for storing x and y values
    x_values = [
        x0 + i * h for i in range(n)
    ]  # Fill in the range from x0 to maximum xmax with the appropriate step
    y_values = [0] * n  # For now, fill in the zeros
    y_values[0] = y0  # The value of the first y[0] = y0, which was set by the user

    # Calculate the value of y using the Euler method
    for i in range(1, n):
        y_values[i] = y_values[i - 1] + h * f(x_values[i - 1], y_values[i - 1])

    # Print x and y in a table in the console
    print('\nEuler\'s method:')
    print("   x\t   y")
    for x, y in zip(x_values, y_values):
        print(f"{x:.2f}\t{y:.6f}")

    return x_values, y_values


# Method for calculating the solution of a differential equation by the Runge-Kutta method
# The method receives a function f, x0 (initial value of x), y0 (initial value of y), h (step) and xmax
# (the maximum value of x up to which the calculation will be performed)
def runge_kutta_method(f, x0, y0, h, xmax):
    # Calculate the number of steps
    n = int((xmax - x0) / h) + 1

    # Initialize arrays for storing x and y values
    x_values = [
        x0 + i * h for i in range(n)
    ]  # Fill in the range from x0 to maximum xmax with the appropriate step
    y_values = [0] * n  # For now, fill in the zeros
    y_values[0] = y0  # The value of the first y[0] = y0, which was set by the user

    # Calculate the value of y using the Runge-Kutta method
    # (first, calculate the coefficients, then calculate y using the coefficients' base values)
    for i in range(1, n):
        k1 = f(x_values[i - 1], y_values[i - 1])
        k2 = f(x_values[i - 1] + 0.5 * h, y_values[i - 1] + 0.5 * h * k1)
        k3 = f(x_values[i - 1] + 0.5 * h, y_values[i - 1] + 0.5 * h * k2)
        k4 = f(x_values[i - 1] + h, y_values[i - 1] + h * k3)

        y_values[i] = y_values[i - 1] + (h / 6) * (k1 + 2 * k2 + 2 * k3 + k4)

    # Tabular output of x and y in the console
    print('\nRunge-Kutta method:')
    print("   x\t   y")
    for x, y in zip(x_values, y_values):
        print(f"{x:.2f}\t{y:.6f}")

    return x_values, y_values


def plot_methods(x_values_euler, y_values_euler, x_values_rk, y_values_rk):
    plt.plot(x_values_euler, y_values_euler, marker='o', label='Euler Method')
    plt.plot(x_values_rk, y_values_rk, marker='o', label='Runge-Kutta Method')
    plt.xlabel('x')
    plt.ylabel('y')
    plt.title('Solutions using Euler and Runge-Kutta Methods')
    plt.legend()
    plt.grid(True)
    plt.show()


def save_results_to_excel(x_values_euler, y_values_euler, x_values_rk, y_values_rk):
    wb = Workbook()
    ws = wb.active

    ws.merge_cells('A1:C1')
    ws['A1'] = 'Euler\'s method'
    ws['A1'].font = Font(bold=True)
    ws['A2'] = 'X'
    ws['B2'] = 'Y'
    ws['A2'].font = Font(bold=True)
    ws['B2'].font = Font(bold=True)
    ws['A2'].alignment = Alignment(horizontal='right')
    ws['B2'].alignment = Alignment(horizontal='right')

    for i, (x, y) in enumerate(zip(x_values_euler, y_values_euler), start=3):
        ws[f"A{i}"] = x
        ws[f"B{i}"] = y

    ws.merge_cells('E1:G1')
    ws['E1'] = 'Runge-Kutta method'
    ws['E1'].font = Font(bold=True)
    ws['E2'] = 'X'
    ws['F2'] = 'Y'
    ws['E2'].font = Font(bold=True)
    ws['F2'].font = Font(bold=True)
    ws['E2'].alignment = Alignment(horizontal='right')
    ws['F2'].alignment = Alignment(horizontal='right')

    for i, (x, y) in enumerate(zip(x_values_rk, y_values_rk), start=3):
        ws[f"E{i}"] = x
        ws[f"F{i}"] = y

    wb.save("euler_and_runge_kutta_results.xlsx")


x0 = 0
y0 = float(input("Enter the initial value of y(0): "))
h = float(input("Enter the value of step h: "))
x_max = float(input("Enter the maximum value of x: "))

x_values_euler, y_values_euler = euler_method(f, x0, y0, h, x_max)
x_values_rk, y_values_rk = runge_kutta_method(f, x0, y0, h, x_max)

plot_methods(x_values_euler, y_values_euler, x_values_rk, y_values_rk)
save_results_to_excel(x_values_euler, y_values_euler, x_values_rk, y_values_rk)
